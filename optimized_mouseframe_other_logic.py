#!/usr/bin/env python3
"""
Optimized version of mouseframe_tiago_logic.py.
Handles Excel data extraction and processing with improved performance and Pythonic standards.
"""

import logging
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# --- Configuration & Constants ---

# Target labels to extract from Excel
TARGET_LABELS = {
    "Stride length left front average in cm:",
    "Stride length right front average in cm:",
    "Stride length left hind average in cm:",
    "Stride length right hind average in cm:",
    "Overlap left average in cm:",
    "Overlap Right average in cm:",
    "Stride Width Front(L) average in cm:",
    "Stride Width Front(R) average in cm:",
    "Stride Width Hind(L) average in cm:",
    "Stride Width Hind(R) average in cm:",
}

# Clean labels (stripped) for consistent matching
CLEAN_TARGET_LABELS = {s.strip() for s in TARGET_LABELS}

# Mapping for unified output labels
RENAME_MAP = {
    "Stride Width Front(L) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Front(R) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Hind(L) average in cm:": "Stride Width Hind average in cm:",
    "Stride Width Hind(R) average in cm:": "Stride Width Hind average in cm:",
}

CLEAN_RENAME_MAP = {k.strip(): v.strip() for k, v in RENAME_MAP.items()}

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)

def extract_table_data(file_path: Path) -> List[Dict[str, Any]]:
    """
    Parses the Excel file and extracts numeric data associated with 'Table ID' markers.
    Uses a single-pass row iteration for O(N) performance.
    """
    # read_only=True is significantly faster and more memory-efficient
    try:
        # data_only=True ensures we get calculated values, not formulas
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        logging.error(f"Could not load Excel file: {e}")
        return []

    extracted_records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Track active tables by their column index: col_idx -> {table_id, values_dict}
        active_tables: Dict[int, Dict[str, Any]] = {}

        def flush_table(col_idx: int):
            """Helper to compute averages and store the table data."""
            if col_idx in active_tables:
                table_info = active_tables[col_idx]
                record = {
                    "Table ID": table_info["Table ID"],
                    "Source Sheet": sheet_name
                }
                # Calculate mean for each label collected
                for label, vals in table_info["values"].items():
                    # We use np.nanmean or check if vals exists to avoid warnings
                    if vals:
                        record[label] = np.mean(vals)
                    else:
                        record[label] = None
                
                extracted_records.append(record)
                del active_tables[col_idx]

        # Single pass through all rows using iter_rows(values_only=True)
        # This avoids the overhead of creating Cell objects.
        for row in ws.iter_rows(values_only=True):
            for col_idx, cell_value in enumerate(row):
                if cell_value is None:
                    continue

                val_str = str(cell_value).strip()

                # Detect start of a new table
                # Original logic searched for "Table ID" and handled "Table ID:"
                if val_str.startswith("Table ID"):
                    flush_table(col_idx)
                    # Extract ID (handles "Table ID: XXX" or "Table ID XXX")
                    table_id = val_str.replace("Table ID:", "").replace("Table ID", "").strip()
                    active_tables[col_idx] = {
                        "Table ID": table_id,
                        "values": defaultdict(list)
                    }
                
                # Check if this column is part of an active table scan
                elif col_idx in active_tables:
                    if val_str in CLEAN_TARGET_LABELS:
                        # Value is expected in the adjacent column
                        if col_idx + 1 < len(row):
                            raw_value = row[col_idx + 1]
                            try:
                                # Ensure we have a valid number
                                numeric_val = float(raw_value)
                                output_label = CLEAN_RENAME_MAP.get(val_str, val_str)
                                active_tables[col_idx]["values"][output_label].append(numeric_val)
                            except (ValueError, TypeError):
                                # Skip non-numeric values (e.g., text, None)
                                pass

        # Flush any remaining tables at the end of the sheet
        for idx in list(active_tables.keys()):
            flush_table(idx)

    return extracted_records

def process_excel(input_path_str: str) -> Optional[Path]:
    """
    Main logic to process the input file and save the cleaned output.
    """
    input_path = Path(input_path_str).resolve()
    if not input_path.is_file():
        logging.error(f"File not found: {input_path}")
        return None

    logging.info(f"Extracting data from: {input_path.name}")
    
    try:
        data = extract_table_data(input_path)
    except Exception as e:
        logging.error(f"An error occurred during extraction: {e}")
        return None

    if not data:
        logging.warning("No 'Table ID' entries or valid data found. Check file formatting.")
        return None

    # Use Pandas for efficient grouping and averaging
    df = pd.DataFrame(data)
    
    # Generate output path
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"MFrame_clean_output_{timestamp}.xlsx"
    output_path = input_path.parent / output_filename

    logging.info("Averaging data and generating output...")

    try:
        # Use ExcelWriter with the openpyxl engine
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Group by original sheet name to preserve the source structure
            for sheet_name, group_df in df.groupby("Source Sheet"):
                # 1. Drop the helper sheet name column
                # 2. Group by Table ID and average all numeric metrics
                # 3. numeric_only=True ensures we don't try to mean the Table ID if it's not numeric
                final_df = (
                    group_df.drop(columns=["Source Sheet"])
                    .groupby("Table ID", as_index=False)
                    .mean(numeric_only=True)
                )
                # Write to the specific sheet
                final_df.to_excel(writer, sheet_name=str(sheet_name), index=False)
        
        logging.info(f"Successfully saved to: {output_path}")
        return output_path
    except Exception as e:
        logging.error(f"Failed to save output file: {e}")
        return None

def main():
    """CLI Entrypoint."""
    print("\n--- MouseFrame Logic Optimizer ---\n")
    user_input = input("Enter the path to the Excel file: ").strip()
    
    if not user_input:
        logging.error("No input provided.")
        return

    result = process_excel(user_input)
    if result:
        print(f"\n✅ Done! Output file: {result}")
    else:
        print("\n❌ Failed to process the file.")

if __name__ == "__main__":
    main()
