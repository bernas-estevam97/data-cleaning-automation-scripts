import pandas as pd
import win32com.client
import sys
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.formula.translate import Translator
from openpyxl.styles import Color, Fill, Font
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
import win32com.client
import os
import sys
from pathlib import Path
import textwrap
from argparse import ArgumentParser, HelpFormatter
import time
from multiprocessing import Process

class RawFormatter(HelpFormatter):
    def _fill_text(self, text, width, indent):
        return "\n".join([textwrap.fill(line, width) for line in textwrap.indent(textwrap.dedent(text), indent).splitlines()])

program_descripton = f'''
    Excel formatter automation tool

    This is the help screen for excel_modifier_automation v1.0!

    -This function receives the absolute path of a folder to work on given by the user.

    -This folder needs to have excel files in it to work on.

    -WORK IN PROGRESS

    -WORK IN PROGRESS'''

parser = ArgumentParser(description=program_descripton, epilog="""Run python <script_name> -h to see this help screen again.""", formatter_class=RawFormatter)
args = parser.parse_args()

# -----------------Cell styling section:----------------

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# ------------------------------------------------------

def full_excel_automation_function():
    # First function ---> TXT TO EXCEL
    folder_input = input('Which folder has your txt files? ')
    if os.path.isdir(folder_input):
        print('Folder selected: ', folder_input)
    elif folder_input == "":
        print("You didn't input any path. Try again")
        full_excel_automation_function()
    else:
        print('Invalid path input. Try again.')
        full_excel_automation_function()
    
    folder_to_save = input('In which folder do you want to save the excel files (press enter if you want to be the same folder as your txt files)? ')
    if os.path.isdir(folder_to_save):
        print('Folder to save: ', folder_to_save)
    elif folder_to_save == '':   
        folder_to_save = folder_input
        print('Folder is the same as your txt files. --> ', folder_to_save)
    else:
        print('Invalid path input. Try again.')
        full_excel_automation_function()
    txt_files = os.listdir(folder_input)
    txt_to_excel = []
    for file in txt_files:
        if os.path.splitext(file)[1] == '.txt':
            txt_to_excel.append(file)
        else:
            pass
    # print(txt_to_excel)
    if len(txt_to_excel) > 0:
        print(f'There are {len(txt_to_excel)} txt files in your supplied folder.')
        for idx, txt in enumerate(txt_to_excel):
            df = pd.read_csv(os.path.join(folder_input,txt), sep='\t')
            with pd.ExcelWriter(os.path.join(folder_to_save,os.path.splitext(txt)[0]+'.xlsx')) as writer:
                df.to_excel(writer, sheet_name='Sheet 1', index=False)
            print(f'File {txt} converted to excel. {idx+1} of {len(txt_to_excel)} done.') 
        print(f'All {len(txt_to_excel)} txt files have been successfully converted to Excel format.')
    else:
        print('Your folder has no txt files')
        print('Restarting function... Press Ctrl+C to exit at any point')
        full_excel_automation_function()

#def excel_modifier_function():
    # Initial section get excel files from supplied folder
    # folder_input_excel = input('Which folder has your excel files? ')
    # if os.path.isdir(folder_input_excel):
    #     print('Folder selected: ', folder_input_excel)
    # elif folder_input_excel == "":
    #     print("You didn't input any path. Try again")
    #     txt_to_excel_function()
    # else:
    #     print('Invalid path input. Try again.')
    #     txt_to_excel_function()
    print('TXT to EXCEL  function finished\nExcel modifier function started:')
    folder_files = os.listdir(folder_to_save)
    excel_files = []
    for file in folder_files:
        if file.endswith('.xlsx'):
            excel_files.append(file)
        else:
            pass
    
    # def excel_file_processing():
    #try:
    #    excel = win32com.client.Dispatch("Excel.Application")
    #    version = excel.version
    #    print("Excel version:", version)
    #except:
    #    print("You don't have Excel installed in this device.")
    #    sys.exit('Program exited because of non-existant version of MS Excel.')
    for idx, excel_file in enumerate(excel_files):
        # OPEN EACH EXCEL FILE AND MODIFY IT
        # INSTRUCTIONS:
        # - Insert new row at the top of the excel file
        # - Merge cells B1 to G1
        # - Write NAO USAR and mark the cell with new color
        # - Insert 6 new columns at index 8
        # - Copy values from B2 to G2 to H2 to M2
        # - Create the formula: =B3*(4/6) at cell H3
        # - Copy and translate the formula to all the cells in the newly 6 created columns (until last row with data)
        # - Save the modified excel file to the folder
        wb = openpyxl.load_workbook(filename = os.path.join(folder_to_save, excel_file))
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        ws.insert_rows(1,1)
        ws.merge_cells('B1:G1')
        ws['B1'] = 'NAO USAR'
        ws['B1'].alignment = Alignment(horizontal = 'center')
        ws['B1'].fill = PatternFill(fill_type='solid', start_color='FEDC00', end_color='FEDC00')
        ws.insert_cols(8, 6)
        ws['H2'], ws['I2'], ws['J2'], ws['K2'], ws['L2'], ws['M2'] = ws['B2'].value, ws['C2'].value, ws['D2'].value, ws['E2'].value, ws['F2'].value, ws['G2'].value
        ws['H2'].border = ws['I2'].border = ws['J2'].border = ws['K2'].border = ws['L2'].border = ws['M2'].border = thin_border
        ws['H2'].font = ws['I2'].font = ws['J2'].font = ws['K2'].font = ws['L2'].font = ws['M2'].font = Font(bold=True)
        ws['H2'].alignment = ws['I2'].alignment = ws['J2'].alignment = ws['K2'].alignment = ws['L2'].alignment = ws['M2'].alignment = Alignment(horizontal = 'center')
        ws['H3'] = "=B3*(4/6)"
        # ws.column_dimensions['B'].hidden, ws.column_dimensions['C'].hidden, ws.column_dimensions['D'].hidden, ws.column_dimensions['E'].hidden, ws.column_dimensions['F'].hidden, ws.column_dimensions['G'].hidden = True
        # ws.column_dimensions['AI'].hidden, ws.column_dimensions['AJ'].hidden, ws.column_dimensions['AK'].hidden, ws.column_dimensions['AL'].hidden, ws.column_dimensions['AM'].hidden, ws.column_dimensions['AN'].hidden, ws.column_dimensions['AO'].hidden, ws.column_dimensions['AP'].hidden= True
        for col in ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=8, max_col=13):
            for c in col:
                ws[str(c.coordinate)] = Translator("=B3*(4/6)", origin="H3").translate_formula(str(c.coordinate))
                # print(c.coordinate)  
        for colu in ['B', 'C', 'D', 'E', 'F', 'G', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP']:
            ws.column_dimensions[colu].hidden = True      
        # wb.save(os.path.join(folder_input_excel, os.path.splitext(excel_file)[0] + '_modified' + os.path.splitext(excel_file)[1]))
        wb.save(os.path.join(folder_to_save, excel_file))
        
        print(f'Excel file: {excel_file} saved, {idx+1} of {len(excel_files)} done.')
    sys.exit('All files processed to Excel.\nProgram finished\n--------------')



if __name__ == '__main__':
    try:
        start_time = time.time()
        full_excel_automation_function()
        total = time.time() - start_time
        print('All functions completed!')
        print(f'Process finished in {total:.2f} s.')
        # time.sleep(1)
        # excel_modifier_function()
        # print('All functions completed!')
    except KeyboardInterrupt: 
        sys.stderr.write("\n--------------------------\nProgram terminated by user\n")
        
    