import os
import win32com.client
import sys
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.formula.translate import Translator
from openpyxl.styles import Color, Fill, Font
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
import concurrent.futures
import time
from multiprocessing import Pool
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed


#try:
#    excel = win32com.client.Dispatch("Excel.Application")
#    version = excel.version
#    print("Excel version:", version)
#except:
#    print("You don't have Excel installed in this device.")
#    sys.exit('Program exited successfully')




thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def excel_modifier(index_file_total_tuple):
        # OPEN EACH EXCEL FILE AND MODIFY IT
        # INSTRUCTIONS:
        # - Insert new row at the top of the excel file
        # - Merge cells B1 to G1
        # - Write NAO USAR and mark the cell with new color
        # - Insert 6 new columns at index 8
        # - Copy values from B2 to G2 to H2 to M2
        # - Create the formula: =B3*(4/6) at cell H3
        # - Copy and translate the formula to all the cells in the newly 6 created columns (until last row with data)
        # - Save the modified excel file to the folder
    index, file_path, total_files = index_file_total_tuple  # Unpack the tuple (index, file_path)
    try:
        #print(f"Processing file {index + 1} of {total_files}: {file_path}")
        wb = openpyxl.load_workbook(filename = file_path)
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        ws.insert_rows(1,1)
        ws.merge_cells('B1:G1')
        ws['B1'] = 'NAO USAR'
        ws['B1'].alignment = Alignment(horizontal = 'center')
        ws['B1'].fill = PatternFill(fill_type='solid', start_color='FEDC00', end_color='FEDC00')
        ws.insert_cols(8, 6)
        ws['H2'], ws['I2'], ws['J2'], ws['K2'], ws['L2'], ws['M2'] = ws['B2'].value, ws['C2'].value, ws['D2'].value, ws['E2'].value, ws['F2'].value, ws['G2'].value
        ws['H2'].border = ws['I2'].border = ws['J2'].border = ws['K2'].border = ws['L2'].border = ws['M2'].border = thin_border
        ws['H2'].font = ws['I2'].font = ws['J2'].font = ws['K2'].font = ws['L2'].font = ws['M2'].font = Font(bold=True)
        ws['H2'].alignment = ws['I2'].alignment = ws['J2'].alignment = ws['K2'].alignment = ws['L2'].alignment = ws['M2'].alignment = Alignment(horizontal = 'center')
        ws['H3'] = "=B3*(4/6)"
        # ws.column_dimensions['B'].hidden, ws.column_dimensions['C'].hidden, ws.column_dimensions['D'].hidden, ws.column_dimensions['E'].hidden, ws.column_dimensions['F'].hidden, ws.column_dimensions['G'].hidden = True
        # ws.column_dimensions['AI'].hidden, ws.column_dimensions['AJ'].hidden, ws.column_dimensions['AK'].hidden, ws.column_dimensions['AL'].hidden, ws.column_dimensions['AM'].hidden, ws.column_dimensions['AN'].hidden, ws.column_dimensions['AO'].hidden, ws.column_dimensions['AP'].hidden= True
        for col in ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=8, max_col=13):
            for c in col:
                ws[str(c.coordinate)] = Translator("=B3*(4/6)", origin="H3").translate_formula(str(c.coordinate))
                # print(c.coordinate)  
        for colu in ['B', 'C', 'D', 'E', 'F', 'G', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP']:
            ws.column_dimensions[colu].hidden = True      
        # wb.save(os.path.join(folder_input_excel, os.path.splitext(excel_file)[0] + '_modified' + os.path.splitext(excel_file)[1]))
        wb.save(file_path)
        return f"Processed file {index + 1} of {total_files}: {file_path}"
        
    except Exception as e:
        return f"Failed to process {file_path}: {e}"
    
    
def main():
    # Get folder path input from user
    folder_input = input('Which folder has your xlsx files? ')
    if os.path.isdir(folder_input):
        print('Folder selected: ', folder_input)
        #folder_files = os.listdir(folder_input)
        
        
        #for file in folder_files:
        #    if file.endswith('.xlsx'):
        #        excel_files.append(file)
        #    else:
        #        pass
    elif folder_input == "":
        print("You didn't input any path.")
        return
        
    else:
        print('Invalid path input.')
        return
        
    
    
    # Get the list of files in the folder (filtering only xlsx files)
    file_paths = [os.path.join(folder_input, str(p)) for p in os.listdir(folder_input) if p.endswith('.xlsx')]
    #if not file_paths:
    #    print(f"No files found in the folder '{folder_input}'.")
    #    return
    total_files = len(file_paths)
    indexed_files = [(index, file_path, total_files) for index, file_path in enumerate(file_paths)]
    # Number of processes to use (you can set this dynamically)
    num_processes = os.cpu_count()
    # ------------------------------------------- #
    
    # Create a Pool of workers and process files in parallel using map
    #with Pool(num_processes) as pool:
    #    pool.map(excel_modifier, indexed_files)
    # ------------------------------------------- #
    
    #Create a Pool of workers using imap_unordered()
    #with Pool(num_processes) as pool:
        # Use imap_unordered to process files in parallel and get results as they are completed
    #    for _ in pool.imap_unordered(excel_modifier, indexed_files):
    #        pass  # We're just processing the files, so no need to do anything with the results here
    # ------------------------------------------- #
    
    #Create a ProcessPoolExecutor
    with ProcessPoolExecutor(max_workers=num_processes) as executor:
        # Submit all the tasks to the pool
        futures = [executor.submit(excel_modifier, indexed_file) for indexed_file in indexed_files]

        # Use as_completed to get results as they are finished
        for future in as_completed(futures):
            try:
                result = future.result()
                print(result)  # Print the result from the completed task
            except Exception as exc:
                print(f"Generated an exception: {exc}")
    # -------------------------------------------- #            
                
                
                
if __name__ == '__main__':
    start = time.perf_counter()
    main()
    
    finish = time.perf_counter()

    print(f'Excel modifier finished in {round(finish-start, 2)} second(s)')