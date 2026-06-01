import openpyxl

# 1. SETUP
input_file = r"C:\Users\berna\Downloads\dados footprint rodrigo.xlsx"  # Update path if needed
output_file = r"C:\Users\berna\Downloads\dados_footprint_rodrigo_FINAL_FIX.xlsx"

# THRESHOLD: Any number > 500 is considered a formatting error (e.g., 4512 -> 4.512)
THRESHOLD = 500 

# KEYWORDS
# If a cell contains these words, we check the cell to the RIGHT.
label_keywords = ['average in cm', 'average in mm', 'overlap', 'stride']

# If a column HEADER contains these, we fix the whole column.
header_keywords = ['Value', 'cm', 'mm', 'px', 'Stride', 'Length', 'Width', 'Area']
skip_keywords = ['Cage', 'Animal', 'Genótipo', 'Gender', 'Group', 'ID', 'Date', 'Entry', 'Points']

print(f"Opening {input_file}...")

try:
    wb = openpyxl.load_workbook(input_file)

    for sheet in wb.worksheets:
        print(f"Processing sheet: {sheet.title}")
        
        # --- STRATEGY 1: Identify Measurement Columns (Header Scan) ---
        cols_to_fix = set()
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            for col_idx, cell_value in enumerate(row, start=1):
                if isinstance(cell_value, str):
                    if any(k in cell_value for k in header_keywords) and \
                       not any(skip in cell_value for skip in skip_keywords):
                        cols_to_fix.add(col_idx)

        # --- STRATEGY 2: Fix Data (Cell by Cell) ---
        for row in sheet.iter_rows(min_row=1):
            for i, cell in enumerate(row):
                
                # --- NEW LOGIC: Text-to-Number Conversion ---
                # We skip "Table ID" cells or any description text, but try to convert everything else.
                if isinstance(cell.value, str):
                    # Skip the specific Table ID headers
                    if "Table ID" in cell.value:
                        continue
                    
                    # Try to convert string to float (e.g. "7.45" -> 7.45)
                    try:
                        # Replace comma with dot just in case (European format fix)
                        clean_val = cell.value.replace(',', '.')
                        cell.value = float(clean_val)
                    except ValueError:
                        # If it's real text (like "Left Front"), ignore it.
                        pass

                # --- EXISTING LOGIC: Fix Stride/Size Values ---
                fix_needed = False
                
                # Check A: Is this cell in a known measurement column?
                if cell.column in cols_to_fix:
                    fix_needed = True
                
                # Check B: "Neighbor Check" - Is the PREVIOUS cell a label?
                if i > 0:
                    prev_cell = row[i-1]
                    if isinstance(prev_cell.value, str):
                        if any(k in prev_cell.value.lower() for k in label_keywords):
                            fix_needed = True

                # --- APPLY FIX ---
                # Now that cell.value is forced to be a number (if possible), we check the size.
                if fix_needed and isinstance(cell.value, (int, float)):
                    
                    # Logic: If HUGE, divide it.
                    if cell.value > THRESHOLD:
                        cell.value = cell.value / 1000.0
                        cell.number_format = '0.000' # Force 3 decimals
                    
                    # Logic: If SMALL (already fixed), just format it nicely.
                    elif cell.value > 0:
                        cell.number_format = '0.000'

    wb.save(output_file)
    print(f"\nDone! Final fixed file saved as: {output_file}")

except FileNotFoundError:
    print(f"Error: Could not find '{input_file}'. Please check the file path.")
except PermissionError:
    print("Error: The file is open in Excel. Please close it and try again.")