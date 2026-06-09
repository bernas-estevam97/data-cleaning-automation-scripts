import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from collections import defaultdict
import numpy as np
import re
import json

# All possible target labels in Excel
RAW_STRINGS = [
    "Stride length left front average in cm:",
    "Stride length right front average in cm:",
    "Stride length left hind average in cm:",
    "Stride length right hind average in cm:",
    "Overlap left average in cm:",
    "Overlap Right average in cm:",
    "Stride Width Front(L) average in cm:",
    "Stride Width Front(R) average in cm:",
    "Stride Width Hind(L) average in cm:",
    "Stride Width Hind(R) average in cm:"
]

# Map specific labels to unified output names
RENAME_MAP = {
    "Stride Width Front(L) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Front(R) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Hind(L) average in cm:": "Stride Width Hind average in cm:",
    "Stride Width Hind(R) average in cm:": "Stride Width Hind average in cm:"
}

def find_table_data(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    result_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().startswith("Table ID"):
                    table_id = cell.value.strip().replace("Table ID:", "").strip()
                    table_values = defaultdict(list)
                    table_values["Table ID"].append(table_id)

                    col_index = cell.column
                    row_index = cell.row

                    for r in range(row_index + 1, ws.max_row + 1):
                        label_cell = ws.cell(row=r, column=col_index)
                        # FIX 1: Look 4 columns over to grab the 'Value(cm) at 38%' data
                        value_cell = ws.cell(row=r, column=col_index + 4)

                        label = label_cell.value
                        value = value_cell.value

                        if label in RAW_STRINGS and isinstance(value, (int, float)):
                            output_label = RENAME_MAP.get(label, label)
                            table_values[output_label].append(value)

                    table_data = {}
                    for key, values in table_values.items():
                        if key == "Table ID":
                            table_data[key] = values[0]
                        else:
                            table_data[key] = float(np.mean(values)) if values else None

                    result_data.append(table_data)

    return result_data


def run(input_file):
    """
    Main entrypoint for processing an Excel file.
    """
    if not os.path.isfile(input_file):
        raise FileNotFoundError(f"File not found: {input_file}")

    # Extract table data
    data = find_table_data(input_file)
    if not data:
        raise ValueError("No 'Table ID' entries found in the provided file.")

    # Group by W-tag if it exists
    w_groups = defaultdict(list)
    for entry in data:
        table_id = entry.get("Table ID", "")
        w_tag = None
        if table_id.endswith("W"):
            # FIX 2: Split by both underscore and hyphen to correctly isolate '4W'
            parts = re.split(r'[_\-]', table_id)
            for part in parts:
                if re.fullmatch(r"\d+W", part) or re.fullmatch(r"\d+\+\d+W", part):
                    w_tag = part
                    break
        if w_tag:
            w_groups[w_tag].append(entry)
        else:
            w_groups["Unknown"].append(entry)

    # Create output workbook
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(os.path.dirname(input_file), f"MFrame_clean_output_{timestamp}.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    def sort_w_tags(w_tag):
        if w_tag == "Unknown":
            return float('inf')
        try:
            match = re.match(r"(\d+)(?:\+(\d+))?W", w_tag)
            if match:
                base = int(match.group(1))
                extra = int(match.group(2)) if match.group(2) else 0
                return base + extra
        except:
            pass
        return float('inf')

    for w_tag, group_data in sorted(w_groups.items(), key=lambda x: sort_w_tags(x[0])):
        df = pd.DataFrame(group_data)
        
        # ---------------------------------------------------------
        # NEW LOGIC: Handle duplicates and perform mean on columns
        # ---------------------------------------------------------
        df = df.groupby("Table ID", as_index=False).mean()
        
        ws = wb.create_sheet(title=w_tag)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

        headers = list(df.columns)
        try:
            table_id_col_index = headers.index("Table ID") + 1
        except ValueError:
            raise ValueError("Table ID column not found in Excel output.")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            table_id_cell = row[table_id_col_index - 1]
            table_id = str(table_id_cell.value).strip()

    wb.save(output_file)
    return output_file


# CLI usability
if __name__ == "__main__":
    input_file = input("Enter the path to the Excel file: ").strip()
    # Strip quotes just in case the file path was dragged and dropped into the terminal
    input_file = input_file.strip('"').strip("'")
    try:
        output_path = run(input_file)
        print(f"✅ Extracted data has been written to: {output_path}")
    except Exception as e:
        print(f"❌ Error: {e}")