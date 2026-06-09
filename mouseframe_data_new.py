import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from collections import defaultdict
import numpy as np
import re

# All possible target labels in Excel
RAW_STRINGS = [
    "Stride length left front average in cm:",
    "Stride length right front average in cm:",
    "Stride length left hind average in cm:",
    "Stride length right hind average in cm:",
    "Overlap left average in cm:",
    "Overlap Right average in cm:",
    "Stride Width Front(L) average in cm:",
    "Stride Width Front(R) average in cm:",
    "Stride Width Hind(L) average in cm:",
    "Stride Width Hind(R) average in cm:"
]

# Map specific labels to unified output names
RENAME_MAP = {
    "Stride Width Front(L) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Front(R) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Hind(L) average in cm:": "Stride Width Hind average in cm:",
    "Stride Width Hind(R) average in cm:": "Stride Width Hind average in cm:"
}

# Dynamic palettes based on original color preferences
# Format: (Hex Color, Use White Font boolean)
MALE_PALETTE = [
    ("FF003366", True),   # Dark Blue
    ("FFADD8E6", False),  # Light Blue
    ("FF27F5F5", False),  # Cyan
    ("FF808080", True),   # Dark Gray (Fallback 1)
    ("FFFFA500", False)   # Orange (Fallback 2)
]

FEMALE_PALETTE = [
    ("FF8B0000", True),   # Dark Red
    ("FFFFC0CB", False),  # Pink
    ("FFF589C5", False),  # Light Pink
    ("FFA9A9A9", False),  # Light Gray (Fallback 1)
    ("FFFFD700", False)   # Gold (Fallback 2)
]


def find_table_data(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    result_data = []
    THRESHOLD = 500

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_sheet_tag = sheet_name 

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().startswith("Table ID"):
                    
                    raw_table_id = cell.value.strip().replace("Table ID:", "").strip()
                    
                    table_values = defaultdict(list)
                    table_values["Table ID"].append(raw_table_id)
                    table_values["Sheet_Source"].append(current_sheet_tag)

                    col_index = cell.column
                    row_index = cell.row

                    for r in range(row_index + 1, ws.max_row + 1):
                        label_cell = ws.cell(row=r, column=col_index)
                        value_cell = ws.cell(row=r, column=col_index + 1)

                        label = label_cell.value
                        value = value_cell.value

                        if label in RAW_STRINGS and isinstance(value, (int, float)):
                            output_label = RENAME_MAP.get(label, label)
                            
                            if value > THRESHOLD:
                                value = value / 1000.0

                            table_values[output_label].append(value)

                    table_data = {}
                    for key, values in table_values.items():
                        if key in ["Table ID", "Sheet_Source"]:
                            table_data[key] = values[0]
                        else:
                            table_data[key] = float(np.mean(values)) if values else None

                    result_data.append(table_data)

    return result_data


def process_file(input_file, mapping_data, genotype_color_index):
    if not os.path.isfile(input_file):
        raise FileNotFoundError(f"File not found: {input_file}")

    print("Extracting data... (this might take a moment)")
    data = find_table_data(input_file)

    if not data:
        raise ValueError("No 'Table ID' entries found in the provided file.")

    w_groups = defaultdict(list)
    
    for entry in data:
        table_id = entry.get("Table ID", "")
        sheet_source = entry.get("Sheet_Source", "Unknown")
        found_tag = None

        parts = re.split(r'[_\s\-]+', table_id)
        
        for part in parts:
            if re.fullmatch(r"T\d+", part) or re.fullmatch(r"\d+W", part) or re.fullmatch(r"\d+\+\d+W", part):
                found_tag = part
                break
        
        if not found_tag:
             found_tag = sheet_source

        if found_tag:
            w_groups[found_tag].append(entry)
        else:
            w_groups["Unknown"].append(entry)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(os.path.dirname(input_file), f"MFrame_clean_output_{timestamp}.xlsx")

    wb = Workbook()
    wb.remove(wb.active) 
    white_font = Font(color="FFFFFFFF")

    def sort_tags(tag):
        nums = re.findall(r'\d+', tag)
        if nums:
            return int(nums[0])
        return float('inf')

    # Dynamic Color Assigner
    def get_dynamic_fill(genotype, sex):
        g = str(genotype).strip().lower()
        s = str(sex).strip().lower()

        if g not in genotype_color_index:
            return None, False

        # Get the assigned index for this genotype
        idx = genotype_color_index[g]

        # Determine which palette to use based on sex
        if s in ("m", "male", "macho"):
            palette = MALE_PALETTE
        elif s in ("f", "female", "fêmea"):
            palette = FEMALE_PALETTE
        else:
            return None, False

        # Use modulo to safely wrap around if there are more genotypes than palette colors
        color_hex, use_white_text = palette[idx % len(palette)]
        return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid"), use_white_text


    for tag, group_data in sorted(w_groups.items(), key=lambda x: sort_tags(x[0])):
        
        clean_data = []
        for d in group_data:
            d_copy = d.copy()
            d_copy.pop("Sheet_Source", None)
            clean_data.append(d_copy)

        df = pd.DataFrame(clean_data)
        
        cols = ['Table ID'] + [c for c in df.columns if c != 'Table ID']
        df = df[cols]

        ws = wb.create_sheet(title=str(tag))

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, float):
                    cell.number_format = '0.000'

        # Apply coloring if mapping data was provided
        if mapping_data:
            headers = list(df.columns)
            try:
                table_id_col_index = headers.index("Table ID") + 1
            except ValueError:
                continue

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                table_id_cell = row[table_id_col_index - 1]
                table_id = str(table_id_cell.value).strip()

                matched_key = next((k for k in mapping_data if k in table_id), None)

                if matched_key:
                    genotype, sex = mapping_data[matched_key]
                    fill, use_white_font = get_dynamic_fill(genotype, sex)

                    if fill:
                        for cell in row:
                            cell.fill = fill
                            if use_white_font:
                                cell.font = white_font

    wb.save(output_file)
    return output_file


if __name__ == "__main__":
    input_file = input("Enter the path to the Excel data file: ").strip().strip('"').strip("'")
    
    print("\n--- Optional: Coloring ---")
    print("If you have a mapping file (Excel or CSV) with 3 columns (ID, Genotype, Sex), paste the path below.")
    print("Press ENTER to skip coloring entirely.")
    map_input = input("Mapping file path (or Enter to skip): ").strip().strip('"').strip("'")
    
    mapping_dict = {}
    genotype_color_index = {}
    
    if map_input:
        try:
            if map_input.lower().endswith('.csv'):
                df_map = pd.read_csv(map_input)
            else:
                df_map = pd.read_excel(map_input)
            
            # Auto-detect unique genotypes to assign colors dynamically
            unique_genotypes = df_map.iloc[:, 1].astype(str).str.strip().str.lower().unique()
            for i, geno in enumerate(unique_genotypes):
                genotype_color_index[geno] = i
                
            for _, row in df_map.iterrows():
                key = str(row.iloc[0]).strip()
                geno = str(row.iloc[1]).strip()
                sx = str(row.iloc[2]).strip()
                mapping_dict[key] = (geno, sx)
                
            print(f"✅ Loaded mapping for {len(mapping_dict)} subjects.")
            print(f"   Detected {len(unique_genotypes)} unique groups.")
            
        except Exception as e:
            print(f"❌ Could not load mapping file: {e}")
            print("Proceeding without coloring...")
            mapping_dict = {}
            genotype_color_index = {}
    
    try:
        output_path = process_file(input_file, mapping_dict, genotype_color_index)
        print(f"\n✅ Extracted data has been written to: {output_path}")
    except Exception as e:
        print(f"\n❌ Error: {e}")