import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from collections import defaultdict
import numpy as np
import re
from tqdm import tqdm

# --- CONFIGURATION ---

# The list of labels determining the SEARCH and the OUTPUT ORDER
RAW_STRINGS_INPUT = [
    "Stride length left front average in cm",
    "Stride length right front average in cm",
    "Stride length left hind average in cm",
    "Stride length right hind average in cm",
    "Overlap left average in cm",
    "Overlap Right average in cm",
    "Stride Width Front(L) average in cm",
    "Stride Width Front(R) average in cm",
    "Stride Width Hind(L) average in cm",
    "Stride Width Hind(R) average in cm"
]

RENAME_MAP_INPUT = {
    "Stride Width Front(L) average in cm": "Stride Width Front average in cm:",
    "Stride Width Front(R) average in cm": "Stride Width Front average in cm:",
    "Stride Width Hind(L) average in cm": "Stride Width Hind average in cm:",
    "Stride Width Hind(R) average in cm": "Stride Width Hind average in cm:"
}

# --- BUILD LOOKUP MAPS ---
# 1. TARGET_MAP: Maps simplified strings (lowercase, no spaces) -> Output Label
# 2. DESIRED_OUTPUT_ORDER: A list of Output Labels in the exact order of RAW_STRINGS_INPUT

TARGET_MAP = {}
DESIRED_OUTPUT_ORDER = []

def simplify_string(text):
    if not isinstance(text, str): return ""
    # Remove all spaces, punctuation, and make lowercase
    return re.sub(r'[\s:_\-\(\)]', '', text).lower()

seen_labels = set()

for s in RAW_STRINGS_INPUT:
    clean_key = simplify_string(s)
    
    # Check if this label needs renaming
    renamed_label = s
    for k, v in RENAME_MAP_INPUT.items():
        if simplify_string(k) == clean_key:
            renamed_label = v
            break
            
    TARGET_MAP[clean_key] = renamed_label
    
    # Add to sort order list (if not already added)
    if renamed_label not in seen_labels:
        DESIRED_OUTPUT_ORDER.append(renamed_label)
        seen_labels.add(renamed_label)


def load_resilient_workbook(filepath):
    try:
        return load_workbook(filepath, data_only=True)
    except Exception:
        pass 
    
    df = None
    try:
        df = pd.read_csv(filepath, sep=None, engine='python', header=None)
    except Exception:
        try:
            df = pd.read_excel(filepath, header=None)
        except Exception:
            return None 

    if df is not None:
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        return wb
    return None

def find_table_data(workbook_path):
    wb = load_resilient_workbook(workbook_path)
    if wb is None:
        return []

    result_data = []
    ws = wb.active 
    
    filename_full = os.path.basename(workbook_path)
    filename_clean = os.path.splitext(filename_full)[0]

    # Initialize with Table ID (Source File column removed)
    table_values = defaultdict(list)
    table_values["Table ID"].append(filename_clean)
    
    found_data_in_file = False

    # Iterate through every row in the sheet
    for row in ws.iter_rows():
        
        # 1. Reconstruct the full text of the row to handle split labels
        row_strings = [str(c.value) for c in row if c.value and isinstance(c.value, str)]
        full_row_text = "".join(row_strings) 
        
        # 2. Check if this row matches one of our targets
        row_key = simplify_string(full_row_text)
        
        matched_label = None
        
        # Check against our target keys
        for target_key, output_label in TARGET_MAP.items():
            if target_key in row_key:
                matched_label = output_label
                break
        
        # 3. If matched, find the first number in the row
        if matched_label:
            found_val = None
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    found_val = cell.value
                    break
                elif isinstance(cell.value, str):
                    try:
                        found_val = float(cell.value)
                        break
                    except ValueError:
                        continue
            
            if found_val is not None:
                table_values[matched_label].append(found_val)
                found_data_in_file = True

    wb.close()
    
    if not found_data_in_file:
        tqdm.write(f"⚠️  Warning: No matching data found in '{filename_full}'")
        return []

    # Calculate Averages for the file
    table_data = {}
    for key, values in table_values.items():
        if key == "Table ID":
            table_data[key] = values[0]
        else:
            table_data[key] = float(np.mean(values)) if values else None
            
    result_data.append(table_data)
    return result_data


def run(input_folder):
    if not os.path.isdir(input_folder):
        raise FileNotFoundError(f"Folder not found: {input_folder}")

    all_data = []

    files_to_process = [
        f for f in os.listdir(input_folder) 
        if f.endswith(".xlsx") 
        and not f.startswith("MFrame_clean_output")
        and not f.startswith("~$")
    ]

    if not files_to_process:
        raise ValueError("No valid Excel files found.")

    print(f"Found {len(files_to_process)} files. Processing...")

    for filename in tqdm(files_to_process, desc="Progress", unit="file"):
        file_path = os.path.join(input_folder, filename)
        file_data = find_table_data(file_path)
        all_data.extend(file_data)

    if not all_data:
        raise ValueError("No data extracted.")

    # --- GROUPING ---
    w_groups = defaultdict(list)
    for entry in all_data:
        table_id = entry.get("Table ID", "")
        w_tag = None
        
        parts = table_id.split("_")
        for part in parts:
            if re.fullmatch(r"\d+[Ww]", part) or re.fullmatch(r"\d+\+\d+[Ww]", part):
                w_tag = part.upper()
                break
        
        if w_tag:
            w_groups[w_tag].append(entry)
        else:
            w_groups["Unknown"].append(entry)

    # --- OUTPUT ---
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(input_folder, f"MFrame_clean_output_{timestamp}.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    def sort_w_tags(w_tag):
        if w_tag == "Unknown": return float('inf')
        try:
            match = re.match(r"(\d+)(?:\+(\d+))?[Ww]", w_tag, re.IGNORECASE)
            if match:
                return int(match.group(1)) + (int(match.group(2)) if match.group(2) else 0)
        except: pass
        return float('inf')

    print("Writing output file...")
    for w_tag, group_data in sorted(w_groups.items(), key=lambda x: sort_w_tags(x[0])):
        df = pd.DataFrame(group_data)
        
        # --- ORDERING COLUMNS ---
        current_cols = set(df.columns)
        
        # Start with Table ID
        final_cols = ["Table ID"] if "Table ID" in current_cols else []
        
        # Append columns in the specific order defined by RAW_STRINGS_INPUT
        for col_name in DESIRED_OUTPUT_ORDER:
            if col_name in current_cols:
                final_cols.append(col_name)
        
        # Append any leftover columns that might have appeared (just in case)
        for col in list(df.columns):
            if col not in final_cols:
                final_cols.append(col)
        
        # Apply the order
        df = df[final_cols]

        ws = wb.create_sheet(title=w_tag)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_file)
    return output_file

if __name__ == "__main__":
    input_folder = input("Enter the path to the folder: ").strip().replace('"', '').replace("'", "")
    try:
        output_path = run(input_folder)
        print(f"\n✅ Done! Output saved to:\n{output_path}")
    except Exception as e:
        print(f"\n❌ Error: {e}")