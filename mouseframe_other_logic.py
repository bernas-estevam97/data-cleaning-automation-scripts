import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from collections import defaultdict
import numpy as np

# All possible target labels in Excel
RAW_STRINGS = [
    "Stride length left front average in cm:",
    "Stride length right front average in cm:",
    "Stride length left hind average in cm:",
    "Stride length right hind average in cm:",
    "Overlap left average in cm:",
    "Overlap Right average in cm:",
    "Stride Width Front(L) average in cm:",
    "Stride Width Front(R) average in cm:",
    "Stride Width Hind(L) average in cm:",
    "Stride Width Hind(R) average in cm:"
]

# Clean the raw strings to prevent mismatching due to trailing spaces
CLEAN_RAW_STRINGS = [s.strip() for s in RAW_STRINGS]

# Map specific labels to unified output names
RENAME_MAP = {
    "Stride Width Front(L) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Front(R) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Hind(L) average in cm:": "Stride Width Hind average in cm:",
    "Stride Width Hind(R) average in cm:": "Stride Width Hind average in cm:"
}

# Clean the rename map keys too
CLEAN_RENAME_MAP = {k.strip(): v.strip() for k, v in RENAME_MAP.items()}

def find_table_data(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    result_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and str(cell.value).strip().startswith("Table ID"):
                    table_id = str(cell.value).strip().replace("Table ID:", "").strip()
                    table_values = defaultdict(list)
                    table_values["Table ID"].append(table_id)
                    table_values["Source Sheet"].append(sheet_name) 

                    col_index = cell.column
                    row_index = cell.row

                    for r in range(row_index + 1, ws.max_row + 1):
                        label_cell = ws.cell(row=r, column=col_index)
                        value_cell = ws.cell(row=r, column=col_index + 1)

                        # NEW LOGIC: Stop scanning if we hit the next table ID to avoid duplicates
                        if isinstance(label_cell.value, str) and str(label_cell.value).strip().startswith("Table ID"):
                            break

                        label = label_cell.value
                        value = value_cell.value

                        if label is not None:
                            clean_label = str(label).strip()
                            # Check against our cleaned list
                            if clean_label in CLEAN_RAW_STRINGS:
                                # NEW LOGIC: Try to cast to float in case Excel stored it as text
                                try:
                                    num_value = float(value)
                                    output_label = CLEAN_RENAME_MAP.get(clean_label, clean_label)
                                    table_values[output_label].append(num_value)
                                except (ValueError, TypeError):
                                    pass # Skip if it truly isn't a number

                    table_data = {}
                    for key, values in table_values.items():
                        if key in ["Table ID", "Source Sheet"]:
                            table_data[key] = values[0]
                        else:
                            table_data[key] = float(np.mean(values)) if values else None

                    result_data.append(table_data)

    return result_data

def run(input_file):
    """
    Main entrypoint for processing an Excel file.
    """
    if not os.path.isfile(input_file):
        raise FileNotFoundError(f"File not found: {input_file}")

    # Extract table data
    data = find_table_data(input_file)
    if not data:
        raise ValueError("No 'Table ID' entries found in the provided file. Check formatting.")

    # Group data by the original Sheet Name
    sheet_groups = defaultdict(list)
    for entry in data:
        sheet_name = entry.pop("Source Sheet") 
        sheet_groups[sheet_name].append(entry)

    # Create output workbook
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(os.path.dirname(input_file), f"MFrame_clean_output_{timestamp}.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    # Iterate through each organized sheet group
    for sheet_name, group_data in sheet_groups.items():
        df = pd.DataFrame(group_data)
        
        # Group identical Table IDs together and averages all other numeric columns.
        df = df.groupby("Table ID", as_index=False).mean()
        
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_file)
    return output_file

# CLI usability
if __name__ == "__main__":
    input_file = input("Enter the path to the Excel file: ").strip()
    try:
        output_path = run(input_file)
        print(f"✅ Extracted data has been written to: {output_path}")
    except Exception as e:
        print(f"❌ Error: {e}")